"""
Enhanced Log Sheet Export Service
Provides PDF, Excel, CSV export functionality with compliance validation
"""

import io
import csv
import json
from typing import List, Dict, Any
from django.http import HttpResponse
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("ReportLab not available. PDF export will be limited.")
    # Define dummy classes for type hints when ReportLab is not available
    class Table:
        pass
    class TableStyle:
        pass
    class Paragraph:
        pass
    class Spacer:
        pass
    class SimpleDocTemplate:
        pass
    class ParagraphStyle:
        pass
    class getSampleStyleSheet:
        pass
    class colors:
        pass
    class A4:
        pass

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("OpenPyXL not available. Excel export will be limited.")


class LogSheetExporter:
    """
    Comprehensive log sheet export service with multiple formats
    """

    def __init__(self, user, log_entries: List, start_date: str, end_date: str):
        self.user = user
        self.log_entries = log_entries
        self.start_date = start_date
        self.end_date = end_date
        self.driver_name = f"{user.first_name} {user.last_name}"

    def export_pdf(self) -> HttpResponse:
        """Generate PDF log sheet with professional formatting"""
        if not REPORTLAB_AVAILABLE:
            return self._fallback_pdf_response()

        try:
            # Create PDF in memory
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Heading1"], fontSize=16, spaceAfter=30, alignment=1  # Center alignment
            )

            # Build content
            story = []

            # Title
            story.append(Paragraph("Driver Log Sheet", title_style))
            story.append(Paragraph(f"Driver: {self.driver_name}", styles["Normal"]))
            story.append(Paragraph(f"Period: {self.start_date} to {self.end_date}", styles["Normal"]))
            story.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
            story.append(Spacer(1, 20))

            # Compliance summary
            compliance_data = self._get_compliance_summary()
            if compliance_data:
                story.append(Paragraph("Compliance Summary", styles["Heading2"]))
                compliance_table = self._create_compliance_table(compliance_data)
                story.append(compliance_table)
                story.append(Spacer(1, 20))

            # Log entries table
            story.append(Paragraph("Log Entries", styles["Heading2"]))
            log_table = self._create_log_table()
            story.append(log_table)

            # Build PDF
            doc.build(story)

            # Prepare response
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="log_sheet_{self.start_date}_to_{self.end_date}.pdf"'
            )

            return response

        except Exception as e:
            logger.error(f"PDF generation error: {str(e)}")
            return self._fallback_pdf_response()

    def export_excel(self) -> HttpResponse:
        """Generate Excel log sheet with formatting and charts"""
        if not OPENPYXL_AVAILABLE:
            return self._fallback_excel_response()

        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Driver Log Sheet"

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Set headers
            headers = [
                "Date",
                "Start Time",
                "End Time",
                "Duration (Hours)",
                "Duty Status",
                "Location",
                "City",
                "State",
                "Remarks",
                "Certified",
                "Compliance Notes",
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Add log entries
            compliance_data = self._get_compliance_summary()
            for row, entry in enumerate(self.log_entries, 2):
                duration = (entry.end_time - entry.start_time).total_seconds() / 3600
                compliance_note = self._get_compliance_note(entry, compliance_data)

                ws.cell(row=row, column=1, value=entry.start_time.date())
                ws.cell(row=row, column=2, value=entry.start_time.time())
                ws.cell(row=row, column=3, value=entry.end_time.time())
                ws.cell(row=row, column=4, value=round(duration, 2))
                ws.cell(row=row, column=5, value=entry.duty_status.name)
                ws.cell(row=row, column=6, value=entry.location)
                ws.cell(row=row, column=7, value=entry.city)
                ws.cell(row=row, column=8, value=entry.state)
                ws.cell(row=row, column=9, value=entry.remarks)
                ws.cell(row=row, column=10, value="Yes" if entry.is_certified else "No")
                ws.cell(row=row, column=11, value=compliance_note)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add summary sheet
            self._add_summary_sheet(wb, compliance_data)

            # Save to memory
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = (
                f'attachment; filename="log_sheet_{self.start_date}_to_{self.end_date}.xlsx"'
            )

            return response

        except Exception as e:
            logger.error(f"Excel generation error: {str(e)}")
            return self._fallback_excel_response()

    def export_csv(self) -> HttpResponse:
        """Generate CSV log sheet with enhanced data"""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="log_sheet_{self.start_date}_to_{self.end_date}.csv"'

        writer = csv.writer(response)

        # Enhanced headers
        headers = [
            "Date",
            "Start Time",
            "End Time",
            "Duration (Hours)",
            "Duty Status",
            "Location",
            "City",
            "State",
            "Remarks",
            "Certified",
            "Driver Name",
            "Driver ID",
            "Export Date",
            "Compliance Status",
        ]
        writer.writerow(headers)

        # Add compliance data
        compliance_data = self._get_compliance_summary()

        for entry in self.log_entries:
            duration = (entry.end_time - entry.start_time).total_seconds() / 3600
            compliance_status = self._get_compliance_status(entry, compliance_data)

            writer.writerow(
                [
                    entry.start_time.date(),
                    entry.start_time.time(),
                    entry.end_time.time(),
                    round(duration, 2),
                    entry.duty_status.name,
                    entry.location,
                    entry.city,
                    entry.state,
                    entry.remarks,
                    "Yes" if entry.is_certified else "No",
                    self.driver_name,
                    self.user.id,
                    timezone.now().date(),
                    compliance_status,
                ]
            )

        return response

    def _get_compliance_summary(self) -> Dict[str, Any]:
        """Get compliance summary for the log entries"""
        try:
            from core_utils.hos_compliance import HOSComplianceEngine, CycleType

            # Get user's cycle type
            try:
                driver_profile = self.user.driverprofile
                cycle_type = (
                    CycleType(driver_profile.cycle_type) if driver_profile.cycle_type else CycleType.SEVENTY_EIGHT
                )
            except:
                cycle_type = CycleType.SEVENTY_EIGHT

            # Convert log entries to compliance engine format
            log_data = []
            for entry in self.log_entries:
                log_data.append(
                    {
                        "id": entry.id,
                        "start_time": entry.start_time,
                        "end_time": entry.end_time,
                        "duty_status": entry.duty_status.name,
                    }
                )

            # Calculate compliance
            engine = HOSComplianceEngine(cycle_type)
            hos_status = engine.calculate_hos_status(log_data)

            return {
                "cycle_type": cycle_type.value,
                "hours_used": hos_status.hours_used_this_cycle,
                "hours_remaining": hos_status.hours_remaining_this_cycle,
                "consecutive_driving_hours": hos_status.consecutive_driving_hours,
                "consecutive_on_duty_hours": hos_status.consecutive_on_duty_hours,
                "consecutive_off_duty_hours": hos_status.consecutive_off_duty_hours,
                "violations": hos_status.violations,
                "is_compliant": len(hos_status.violations) == 0,
            }
        except Exception as e:
            logger.error(f"Compliance calculation error: {str(e)}")
            return {}

    def _create_log_table(self) -> Table:
        """Create formatted log entries table for PDF"""
        data = [["Date", "Start", "End", "Duration", "Status", "Location", "Certified"]]

        for entry in self.log_entries:
            duration = (entry.end_time - entry.start_time).total_seconds() / 3600
            data.append(
                [
                    entry.start_time.strftime("%m/%d/%Y"),
                    entry.start_time.strftime("%H:%M"),
                    entry.end_time.strftime("%H:%M"),
                    f"{duration:.1f}h",
                    entry.duty_status.name,
                    f"{entry.city}, {entry.state}",
                    "Yes" if entry.is_certified else "No",
                ]
            )

        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                ]
            )
        )

        return table

    def _create_compliance_table(self, compliance_data: Dict) -> Table:
        """Create compliance summary table for PDF"""
        data = [
            ["Compliance Metric", "Value", "Status"],
            ["Cycle Type", compliance_data.get("cycle_type", "N/A"), ""],
            ["Hours Used", f"{compliance_data.get('hours_used', 0):.1f}", ""],
            ["Hours Remaining", f"{compliance_data.get('hours_remaining', 0):.1f}", ""],
            ["Consecutive Driving", f"{compliance_data.get('consecutive_driving_hours', 0):.1f}", ""],
            ["Violations", len(compliance_data.get("violations", [])), ""],
            ["Overall Status", "Compliant" if compliance_data.get("is_compliant", False) else "Non-Compliant", ""],
        ]

        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                ]
            )
        )

        return table

    def _add_summary_sheet(self, wb, compliance_data: Dict):
        """Add compliance summary sheet to Excel workbook"""
        ws_summary = wb.create_sheet("Compliance Summary")

        # Summary data
        summary_data = [
            ["Driver Log Sheet Summary", ""],
            ["Driver Name", self.driver_name],
            ["Period", f"{self.start_date} to {self.end_date}"],
            ["Total Entries", len(self.log_entries)],
            ["", ""],
            ["Compliance Metrics", ""],
            ["Cycle Type", compliance_data.get("cycle_type", "N/A")],
            ["Hours Used", f"{compliance_data.get('hours_used', 0):.1f}"],
            ["Hours Remaining", f"{compliance_data.get('hours_remaining', 0):.1f}"],
            ["Consecutive Driving Hours", f"{compliance_data.get('consecutive_driving_hours', 0):.1f}"],
            ["Violations Count", len(compliance_data.get("violations", []))],
            ["Overall Status", "Compliant" if compliance_data.get("is_compliant", False) else "Non-Compliant"],
        ]

        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)

            if label.endswith("Summary") or label == "Compliance Metrics":
                ws_summary.cell(row=row, column=1).font = Font(bold=True, size=14)
            elif label != "":
                ws_summary.cell(row=row, column=1).font = Font(bold=True)

    def _get_compliance_note(self, entry, compliance_data: Dict) -> str:
        """Get compliance note for a specific log entry"""
        violations = compliance_data.get("violations", [])
        for violation in violations:
            if hasattr(violation, "log_entry_id") and violation.log_entry_id == entry.id:
                return violation.description
        return "Compliant"

    def _get_compliance_status(self, entry, compliance_data: Dict) -> str:
        """Get compliance status for a specific log entry"""
        violations = compliance_data.get("violations", [])
        for violation in violations:
            if hasattr(violation, "log_entry_id") and violation.log_entry_id == entry.id:
                return "Violation"
        return "Compliant"

    def _fallback_pdf_response(self) -> HttpResponse:
        """Fallback PDF response when ReportLab is not available"""
        return HttpResponse(
            json.dumps(
                {
                    "error": "PDF generation not available",
                    "message": "ReportLab library is required for PDF export",
                    "suggestion": "Please install ReportLab or use CSV/Excel export",
                }
            ),
            content_type="application/json",
            status=503,
        )

    def _fallback_excel_response(self) -> HttpResponse:
        """Fallback Excel response when OpenPyXL is not available"""
        return HttpResponse(
            json.dumps(
                {
                    "error": "Excel generation not available",
                    "message": "OpenPyXL library is required for Excel export",
                    "suggestion": "Please install OpenPyXL or use CSV export",
                }
            ),
            content_type="application/json",
            status=503,
        )


class LogComplianceValidator:
    """
    Validates log entries for FMCSA compliance
    """

    def __init__(self, user):
        self.user = user

    def validate_logs(self, log_entries: List) -> Dict[str, Any]:
        """Validate a list of log entries for compliance"""
        try:
            from core_utils.hos_compliance import HOSComplianceEngine, CycleType

            # Get user's cycle type
            try:
                driver_profile = self.user.driverprofile
                cycle_type = (
                    CycleType(driver_profile.cycle_type) if driver_profile.cycle_type else CycleType.SEVENTY_EIGHT
                )
            except:
                cycle_type = CycleType.SEVENTY_EIGHT

            # Convert to compliance engine format
            log_data = []
            for entry in log_entries:
                log_data.append(
                    {
                        "id": entry.id,
                        "start_time": entry.start_time,
                        "end_time": entry.end_time,
                        "duty_status": entry.duty_status.name,
                    }
                )

            # Calculate compliance
            engine = HOSComplianceEngine(cycle_type)
            hos_status = engine.calculate_hos_status(log_data)

            return {
                "is_valid": len(hos_status.violations) == 0,
                "violations": hos_status.violations,
                "warnings": self._generate_warnings(hos_status),
                "recommendations": self._generate_recommendations(hos_status),
                "summary": {
                    "total_hours": hos_status.hours_used_this_cycle,
                    "remaining_hours": hos_status.hours_remaining_this_cycle,
                    "consecutive_driving": hos_status.consecutive_driving_hours,
                    "consecutive_on_duty": hos_status.consecutive_on_duty_hours,
                },
            }

        except Exception as e:
            logger.error(f"Compliance validation error: {str(e)}")
            return {"is_valid": False, "error": str(e), "violations": [], "warnings": [], "recommendations": []}

    def _generate_warnings(self, hos_status) -> List[str]:
        """Generate warnings based on HOS status"""
        warnings = []

        if hos_status.hours_remaining_this_cycle < 2:
            warnings.append("Less than 2 hours remaining in current cycle")

        if hos_status.consecutive_driving_hours > 9:
            warnings.append("Approaching 11-hour driving limit")

        if hos_status.consecutive_on_duty_hours > 12:
            warnings.append("Approaching 14-hour on-duty limit")

        return warnings

    def _generate_recommendations(self, hos_status) -> List[str]:
        """Generate recommendations based on HOS status"""
        recommendations = []

        if hos_status.hours_remaining_this_cycle < 4:
            recommendations.append("Consider taking a 34-hour restart soon")

        if hos_status.consecutive_driving_hours > 8:
            recommendations.append("Plan for a 30-minute break within 3 hours")

        if hos_status.consecutive_on_duty_hours > 10:
            recommendations.append("Plan for 10-hour off-duty period soon")

        return recommendations
